# %%
import warnings as _warnings
import pandas as _pd
import openpyxl as _openpyxl
import os as _os
import fnmatch as _fnmatch
import copy as _copy

from dataclasses import dataclass as _dataclass
from openpyxl.worksheet._read_only import ReadOnlyWorksheet as _WS
from typing import Callable as _Callable, Literal as _Literal


@_dataclass
class Cell:
    value: str
    row_index: int
    column_index: int


class Data:
    def __init__(self, data: list[list]):
        self.__data = data if data else []

    def __set__(self, value):
        self.__data = value

    def __iter__(self):
        return iter(self.__data)

    def __repr__(self):
        return "\n".join(map(str, self.__data))

    def __len__(self):
        return len(self.__data)

    def remove_empty_rows(self):
        return Data(
            [
                row
                for row in self.__data
                if any(cell is not None and cell != 0 for cell in row)
            ]
        )

    def remove_empty_columns(self):
        transposed_worksheet = list(map(list, zip(*self.__data)))
        cleaned_transposed_worksheet = [
            col for col in transposed_worksheet if any(cell is not None for cell in col)
        ]
        cleaned_worksheet = list(map(list, zip(*cleaned_transposed_worksheet)))

        return Data(cleaned_worksheet)

    def remove_empty_rows_and_columns(self):
        return self.remove_empty_rows().remove_empty_columns()

    def clean_multi_row_columns(self, height: int):
        if height == 0 or height == 1:
            return Data(self.__data[0]) if len(self.__data) > 0 else Data([])

        new_columns = []
        for p_index, row in enumerate(self.__data[0:height]):
            new_array = []
            last_value = ""
            for c_index, item in enumerate(row):
                if item is None:
                    if c_index == 0:
                        new_array.append(self.__data[0:height][p_index - 1][c_index])
                        continue
                    if c_index > 0 and last_value:
                        new_array.append(last_value)
                else:
                    new_array.append(item)
                    last_value = item

            new_columns.append(new_array)

        final = [*new_columns, *self.__data[height:]]

        return Data(final)

    def to_dataframe(self, header_row_height: int = 1):
        return _pd.DataFrame(
            self.__data[header_row_height:], columns=self.__data[0:header_row_height]
        )

    def set_bounds(
        self,
        top_left_callback: _Callable[[Cell], bool],
        top_right_callback: _Callable[[Cell], bool] = None,
        bottom_left_callback: _Callable[[Cell], bool] = None,
        bottom_right_callback: _Callable[[Cell], bool] = None,
    ):
        return Bounds(self.__data).set_bounds(
            top_left_callback,
            top_right_callback,
            bottom_left_callback,
            bottom_right_callback,
        )


class Bounds:
    def __init__(self, data: Data):
        self.__data = data if data else []
        self.__bounds = []
        self.__max_rows = len(self.__data)
        self.__max_columns = len(self.__data[0])
        self.__top_row_offset = 0
        self.__right_column_offset = 0
        self.__bottom_row_offset = 0
        self.__left_column_offset = 0

    def __get_coordinaets(
        self, callback: _Callable[[Cell], bool] | None, start: int = 0
    ):
        if callback is None:
            return []

        coordinates = []
        for r_index, row in enumerate(self.__data[start:]):
            for c_index, cell in enumerate(row):
                if callback(Cell(cell, r_index + start, c_index)):
                    coordinates.append((r_index + start, c_index))
        return coordinates

    def __safe_get(self, array: list, index: int):
        try:
            return array[index]
        except IndexError:
            return None

    def __safe_bound(
        self, side: _Literal["top", "bottom", "left", "right"], value: int
    ):
        bounds = {
            "top": (0, self.__max_rows),
            "bottom": (0, self.__max_rows),
            "left": (0, self.__max_columns),
            "right": (0, self.__max_columns),
        }

        min_bound, max_bound = bounds[side]
        clipped_value = min(max(value, min_bound), max_bound)

        if clipped_value != value:
            _warnings.warn(
                f"Value {value} for side {side} is out of bounds. Clipping to {clipped_value}.",
                UserWarning,
                stacklevel=5,
            )

        return clipped_value

    def __get_fallbacks(
        self,
        index: int,
        top_left_c: list[tuple[int, int]],
        top_left: tuple[int, int],
        top_right: tuple[int, int] | None,
        bottom_left: tuple[int, int] | None,
        bottom_right: tuple[int, int] | None,
    ):
        if bottom_left is None:
            for index_c, tl in enumerate(top_left_c, start=index):
                if tl[1] == top_left[1] and tl[0] > top_left[0]:
                    bottom_left = (min(tl[0] - 1, self.__max_rows), top_left[1])
                    break
            else:
                bottom_left = (self.__max_rows, top_left[1])

        if top_right is None:
            for index_c, tl in enumerate(top_left_c, start=index):
                if tl[0] == top_left[0] and tl[1] > top_left[1]:
                    top_right = (top_left[0], min(tl[1] - 1, self.__max_columns))
                    break
            else:
                top_right = (top_left[0], self.__max_columns)

        if bottom_right is None:
            bottom_right = (bottom_left[0], top_right[1])

        return (top_left, top_right, bottom_left, bottom_right)

    def set_bounds(
        self,
        top_left_callback: _Callable[[Cell], bool],
        top_right_callback: _Callable[[Cell], bool] = None,
        bottom_left_callback: _Callable[[Cell], bool] = None,
        bottom_right_callback: _Callable[[Cell], bool] = None,
    ):
        new_bounds = _copy.deepcopy(self)

        top_left_c = self.__get_coordinaets(top_left_callback)
        first_row = self.__safe_get(self.__safe_get(top_left_c, 0), 0) or 0

        top_right_c = self.__get_coordinaets(top_right_callback, start=first_row)
        bottom_left_c = self.__get_coordinaets(bottom_left_callback, start=first_row)
        bottom_right_c = self.__get_coordinaets(bottom_right_callback, start=first_row)

        for index, top_left in enumerate(top_left_c):
            coords = self.__get_fallbacks(
                index=index,
                top_left_c=top_left_c,
                top_left=top_left,
                top_right=self.__safe_get(top_right_c, index),
                bottom_left=self.__safe_get(bottom_left_c, index),
                bottom_right=self.__safe_get(bottom_right_c, index),
            )

            new_bounds.__bounds.append(coords)

        return new_bounds

    def offset(
        self,
        top_row: int = 0,
        left_column: int = 0,
        bottom_row: int = 0,
        right_column: int = 0,
    ):
        new_bounds = _copy.deepcopy(self)
        new_bounds.__top_row_offset = top_row
        new_bounds.__right_column_offset = right_column
        new_bounds.__bottom_row_offset = bottom_row
        new_bounds.__left_column_offset = left_column

        return new_bounds

    def set_max_bounds(self, max_rows: int = None, max_columns: int = None):
        new_bounds = _copy.deepcopy(self)
        new_bounds.__max_rows = max_rows or self.__max_rows
        new_bounds.__max_columns = max_columns or self.__max_columns

        return new_bounds

    def generate_datasets(self):
        for index, _ in enumerate(self.__bounds):
            self.__bounds[index] = (
                (
                    self.__safe_bound(
                        "top", self.__bounds[index][0][0] + self.__top_row_offset
                    ),
                    self.__safe_bound(
                        "left", self.__bounds[index][0][1] + self.__left_column_offset
                    ),
                ),
                (
                    self.__safe_bound(
                        "top", self.__bounds[index][1][0] + self.__top_row_offset
                    ),
                    self.__safe_bound(
                        "right", self.__bounds[index][1][1] + self.__right_column_offset
                    ),
                ),
                (
                    self.__safe_bound(
                        "bottom", self.__bounds[index][2][0] + self.__bottom_row_offset
                    ),
                    self.__safe_bound(
                        "left", self.__bounds[index][2][1] + self.__left_column_offset
                    ),
                ),
                (
                    self.__safe_bound(
                        "bottom", self.__bounds[index][3][0] + self.__bottom_row_offset
                    ),
                    self.__safe_bound(
                        "right", self.__bounds[index][3][1] + self.__right_column_offset
                    ),
                ),
            )

        data: list[Data] = []
        for bound in self.__bounds:
            data.append(
                Data(
                    [
                        [
                            self.__data[r_index][c_index]
                            for c_index in range(bound[0][1], bound[1][1])
                        ]
                        for r_index in range(bound[0][0], bound[2][0] + 1)
                    ]
                )
            )

        return data


class Worksheet:
    def __init__(self, title: str):
        self.title = title
        self.data: Data = None

    def load_raw_data(self, data: _WS):
        self.data = Data([row for row in data.iter_rows(values_only=True)])
        return self

    def load_data(self, data: list[list]):
        self.data = Data(data)
        return self


class Workbook:
    def __init__(self, file_name: str, ws: list[_WS], alias: str = ""):
        self.file_name = file_name
        self.alias = alias
        self.worksheets: list[Worksheet] = []
        self.ws_count = 0

        for sheet in ws:
            if type(sheet) is not _WS:
                raise Exception("Invalid worksheet type")

            self.worksheets.append(Worksheet(title=sheet.title).load_raw_data(sheet))

        self.ws_count = len(self.worksheets)

    def append_sheet(self, title: str, data: list[list]):
        self.worksheets.append(Worksheet(title=title).load_data(data))
        self.ws_count = len(self.worksheets)
        return self

    def remove_sheet(self, *, title: str = None, index: int = None):
        if title is not None:
            self.worksheets = [
                sheet for sheet in self.worksheets if sheet.title != title
            ]
        elif index is not None:
            self.worksheets.pop(index)
        else:
            raise Exception("Either title or index must be provided")

        return self


class Parser:
    def __init__(self):
        self.workbooks: list[Workbook] = []
        self.wb_count = 0
        self.index = 0
        pass

    def __get_file_data(self, file_path: str) -> Workbook:
        self.index = self.index + 1
        wb = _openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
        return Workbook(file_path, wb.worksheets, f"T{self.index}")

    def __reload_parser(self, workbooks: list[Workbook]):
        self.workbooks = _copy.deepcopy(workbooks)
        self.wb_count = len(self.workbooks)
        self.index = self.wb_count
        return self

    def load_files(
        self,
        folder_path: str,
        include_subfolders: bool,
        file_types: list[str] = ["xlsx", "csv"],
        workbooks_to_ignore: list[str] = [],
        workbooks_to_include: list[str] = [],
    ):
        if not _os.path.exists(folder_path):
            raise Exception("Folder not found")

        def __is_valid_file(
            file_path: str,
            file_types: list[str],
            keep: list[str] = None,
            ignore: list[str] = None,
        ) -> bool:
            for file_type in file_types:
                if _fnmatch.fnmatch(
                    file_path, f"*{file_type}"
                ) and not _fnmatch.fnmatch(file_path, "*~$*"):
                    if keep:
                        for item in keep:
                            if _fnmatch.fnmatch(file_path, f"*{item}*"):
                                return True

                    if ignore:
                        for item in ignore:
                            if _fnmatch.fnmatch(file_path, f"*{item}*"):
                                return False

                    return True
            return False

        if include_subfolders:
            for path, subdirs, files in _os.walk(folder_path):
                for name in files:
                    file_path = _os.path.join(path, name)

                    if __is_valid_file(
                        file_path, file_types, workbooks_to_include, workbooks_to_ignore
                    ):
                        file_data = self.__get_file_data(file_path)
                        self.workbooks.append(file_data)
        else:
            for file in _os.listdir(folder_path):
                file_path = _os.path.join(folder_path, file)

                if __is_valid_file(
                    file_path, file_types, workbooks_to_include, workbooks_to_ignore
                ):
                    file_data = self.__get_file_data(file_path)
                    self.workbooks.append(file_data)

        self.wb_count = len(self.workbooks)

        return self

    def load_file(self, file_path: str):
        if not _os.path.exists(file_path):
            raise Exception("File not found")
        else:
            self.workbooks.append(self.__get_file_data(file_path))

        self.wb_count = len(self.workbooks)

        return self

    def pipe_workbooks(
        self, callback: _Callable[[Workbook, list[Workbook]], Workbook | None]
    ):
        new_parser = Parser().__reload_parser(self.workbooks)
        for index, wb in enumerate(new_parser.workbooks.copy()):
            new_wb = callback(wb, new_parser.workbooks)
            if new_wb is not None:
                new_parser.workbooks[index] = new_wb

        return new_parser
